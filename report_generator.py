import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from calendar import month_abbr
from datetime import datetime

# ── Color constants ──
DARK_BLUE_FILL = PatternFill("solid", fgColor="2F5496")
BLUE_FILL = PatternFill("solid", fgColor="4472C4")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
GRAY_FILL = PatternFill("solid", fgColor="D6DCE4")
WHITE_FONT = Font(name="Arial", bold=True, color="FFFFFF")
WHITE_FONT_12 = Font(name="Arial", bold=True, color="FFFFFF", size=12)
WHITE_FONT_14 = Font(name="Arial", bold=True, color="FFFFFF", size=14)
BLUE_FONT_12 = Font(name="Arial", bold=True, color="4472C4", size=12)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
NORMAL_FONT = Font(name="Arial", size=10)
GREEN_FONT = Font(name="Arial", size=10, color="006100")
YELLOW_FONT = Font(name="Arial", size=10, color="9C6500")
RED_FONT = Font(name="Arial", size=10, color="9C0006")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def categorize(mins):
    if mins <= 15:
        return "Within 15 mins"
    elif mins <= 30:
        return "16-30 min"
    else:
        return "Over 30 min"


CAT_ORDER = ["Within 15 mins", "16-30 min", "Over 30 min"]


def cat_fill(cat):
    return {
        "Within 15 mins": GREEN_FILL,
        "16-30 min": YELLOW_FILL,
        "Over 30 min": RED_FILL,
    }.get(cat)


def cat_font(cat):
    return {
        "Within 15 mins": GREEN_FONT,
        "16-30 min": YELLOW_FONT,
        "Over 30 min": RED_FONT,
    }.get(cat, NORMAL_FONT)


def style_cell(cell, font=None, fill=None, alignment=None, border=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border


def apply_border_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = THIN_BORDER


# ── Data loading ──
def load_attendance(filepath):
    df = pd.read_excel(filepath, sheet_name=0, header=None)

    # Find header row containing "Employee Code"
    header_row = None
    for i in range(min(20, len(df))):
        row_vals = df.iloc[i].astype(str).str.strip().tolist()
        if "Employee Code" in row_vals:
            header_row = i
            break
    if header_row is None:
        raise ValueError("Cannot find header row with 'Employee Code' in the file.")

    # Re-read with proper header
    df = pd.read_excel(filepath, sheet_name=0, header=header_row)
    df.columns = df.columns.str.strip()

    # Identify key columns (handle possible variations)
    col_map = {}
    for col in df.columns:
        cl = col.lower().replace(" ", "")
        if "employeecode" in cl:
            col_map["code"] = col
        elif "employeename" in cl:
            col_map["name"] = col
        elif "clockdate" in cl:
            col_map["date"] = col
        elif "shiftcode" in cl:
            col_map["shift"] = col
        elif "clockin" == cl or "clockin" in cl:
            if "clock" in cl and "in" in cl and "out" not in cl:
                col_map["clock_in"] = col
        elif "clockout" in cl:
            col_map["clock_out"] = col
        elif "latenessactual" in cl:
            col_map["lateness"] = col
        elif "latenesscount" in cl:
            col_map["late_count"] = col

    # Handle Clock In column more carefully
    if "clock_in" not in col_map:
        for col in df.columns:
            if col.strip().lower() == "clock in":
                col_map["clock_in"] = col
                break

    required = ["code", "name", "date", "lateness"]
    for r in required:
        if r not in col_map:
            raise ValueError(f"Cannot find required column: {r}. Available: {list(df.columns)}")

    # Forward-fill employee code and name
    df[col_map["code"]] = df[col_map["code"]].ffill()
    df[col_map["name"]] = df[col_map["name"]].ffill()

    # Drop rows without valid date
    df = df.dropna(subset=[col_map["date"]])
    df[col_map["date"]] = pd.to_datetime(df[col_map["date"]], errors="coerce")
    df = df.dropna(subset=[col_map["date"]])

    # Convert lateness to numeric
    df[col_map["lateness"]] = pd.to_numeric(df[col_map["lateness"]], errors="coerce").fillna(0)

    # Filter to late records only
    late_df = df[df[col_map["lateness"]] > 0].copy()
    late_df["late_mins"] = (late_df[col_map["lateness"]] * 60).round().astype(int)
    late_df["category"] = late_df["late_mins"].apply(categorize)
    late_df["month_dt"] = late_df[col_map["date"]].dt.to_period("M")
    late_df["month_label"] = late_df[col_map["date"]].dt.strftime("%b %Y")
    late_df["day_name"] = late_df[col_map["date"]].dt.strftime("%a")

    # Standardize column names for downstream
    result = pd.DataFrame({
        "emp_code": late_df[col_map["code"]].astype(str).str.strip(),
        "emp_name": late_df[col_map["name"]].astype(str).str.strip(),
        "date": late_df[col_map["date"]],
        "clock_in": late_df[col_map.get("clock_in", col_map["date"])].astype(str).str.strip() if "clock_in" in col_map else "",
        "late_mins": late_df["late_mins"],
        "category": late_df["category"],
        "month_dt": late_df["month_dt"],
        "month_label": late_df["month_label"],
        "day_name": late_df["day_name"],
    })
    result = result.sort_values(["emp_code", "date"]).reset_index(drop=True)

    # Clean clock_in: take only time part
    def clean_time(v):
        v = str(v).strip()
        if v in ("nan", "", "NaT"):
            return ""
        if " " in v:
            v = v.split(" ")[-1]
        # Remove seconds if present
        parts = v.split(":")
        if len(parts) >= 2:
            return f"{parts[0]}:{parts[1]}"
        return v
    result["clock_in"] = result["clock_in"].apply(clean_time)

    return result


def get_sorted_months(df):
    months = df["month_dt"].unique()
    months = sorted(months)
    labels = []
    for m in months:
        ts = m.to_timestamp()
        labels.append(ts.strftime("%b %Y"))
    return months, labels


# ── Sheet builders ──

def build_summary_mins(wb, df, months, month_labels, top_n):
    ws = wb.create_sheet("Summary of Late in Mins")

    # Title
    ws.merge_cells("A1:R1")
    c = ws["A1"]
    c.value = "Summary of Late in Mins"
    style_cell(c, Font(name="Arial", bold=True, size=14), alignment=CENTER)

    col = 1
    for idx, (m, label) in enumerate(zip(months, month_labels)):
        # Month header
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
        c = ws.cell(2, col, label)
        style_cell(c, BLUE_FONT_12, alignment=CENTER)

        # Column headers
        for j, hdr in enumerate(["Rank", "Staff Name", "Total Mins"]):
            c = ws.cell(3, col + j, hdr)
            style_cell(c, WHITE_FONT, BLUE_FILL, CENTER, THIN_BORDER)

        # Monthly data: sum of late_mins per employee
        m_df = df[df["month_dt"] == m]
        agg = m_df.groupby("emp_name")["late_mins"].sum().reset_index()
        agg = agg.sort_values("late_mins", ascending=False).head(top_n).reset_index(drop=True)

        for i, row in agg.iterrows():
            r = 4 + i
            ws.cell(r, col, i + 1).alignment = CENTER
            ws.cell(r, col + 1, row["emp_name"]).alignment = LEFT
            ws.cell(r, col + 2, row["late_mins"]).alignment = CENTER
            for j in range(3):
                ws.cell(r, col + j).border = THIN_BORDER
                ws.cell(r, col + j).font = NORMAL_FONT

        # Column widths
        ws.column_dimensions[get_column_letter(col)].width = 6
        ws.column_dimensions[get_column_letter(col + 1)].width = 28
        ws.column_dimensions[get_column_letter(col + 2)].width = 12

        col += 3
        # Separator column
        if idx < len(months) - 1:
            ws.column_dimensions[get_column_letter(col)].width = 2
            col += 1


def build_summary_freq(wb, df, months, month_labels, top_n):
    ws = wb.create_sheet("Summary of Late in Freq.")

    ws.merge_cells("A1:R1")
    c = ws["A1"]
    c.value = "Summary of Late in Freq."
    style_cell(c, Font(name="Arial", bold=True, size=14), alignment=CENTER)

    col = 1
    for idx, (m, label) in enumerate(zip(months, month_labels)):
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
        c = ws.cell(2, col, label)
        style_cell(c, BLUE_FONT_12, alignment=CENTER)

        for j, hdr in enumerate(["Rank", "Staff Name", "Frequency"]):
            c = ws.cell(3, col + j, hdr)
            style_cell(c, WHITE_FONT, BLUE_FILL, CENTER, THIN_BORDER)

        m_df = df[df["month_dt"] == m]
        agg = m_df.groupby("emp_name").size().reset_index(name="freq")
        agg = agg.sort_values("freq", ascending=False).head(top_n).reset_index(drop=True)

        for i, row in agg.iterrows():
            r = 4 + i
            ws.cell(r, col, i + 1).alignment = CENTER
            ws.cell(r, col + 1, row["emp_name"]).alignment = LEFT
            ws.cell(r, col + 2, row["freq"]).alignment = CENTER
            for j in range(3):
                ws.cell(r, col + j).border = THIN_BORDER
                ws.cell(r, col + j).font = NORMAL_FONT

        ws.column_dimensions[get_column_letter(col)].width = 6
        ws.column_dimensions[get_column_letter(col + 1)].width = 28
        ws.column_dimensions[get_column_letter(col + 2)].width = 12

        col += 3
        if idx < len(months) - 1:
            ws.column_dimensions[get_column_letter(col)].width = 2
            col += 1


def build_monthly_summary(wb, df, months, month_labels):
    ws = wb.create_sheet("Monthly Summary")

    # Header row 1: Employee Name + month groups
    ws.cell(1, 1, "Employee Name")
    style_cell(ws.cell(1, 1), WHITE_FONT, DARK_BLUE_FILL, CENTER, THIN_BORDER)
    ws.cell(2, 1, "")
    style_cell(ws.cell(2, 1), BOLD_FONT, DARK_BLUE_FILL, CENTER, THIN_BORDER)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.column_dimensions["A"].width = 28

    col = 2
    sub_headers = ["Within 15 mins", "16-30 min", "Over 30 min", "Total"]
    sub_fills = [GREEN_FILL, YELLOW_FILL, RED_FILL, GRAY_FILL]

    for idx, label in enumerate(month_labels):
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
        c = ws.cell(1, col, label)
        style_cell(c, WHITE_FONT, DARK_BLUE_FILL, CENTER, THIN_BORDER)
        for j in range(4):
            c2 = ws.cell(2, col + j, sub_headers[j])
            style_cell(c2, BOLD_FONT, sub_fills[j], CENTER, THIN_BORDER)
            ws.column_dimensions[get_column_letter(col + j)].width = 14
        col += 4

    # Pivot: employee x month x category
    employees = sorted(df["emp_name"].unique())
    pivot = df.groupby(["emp_name", "month_dt", "category"]).size().reset_index(name="count")

    for i, emp in enumerate(employees):
        r = 3 + i
        ws.cell(r, 1, emp)
        style_cell(ws.cell(r, 1), NORMAL_FONT, alignment=LEFT, border=THIN_BORDER)

        col = 2
        for m in months:
            total = 0
            for j, cat in enumerate(CAT_ORDER):
                val = pivot[(pivot["emp_name"] == emp) & (pivot["month_dt"] == m) & (pivot["category"] == cat)]
                cnt = int(val["count"].values[0]) if len(val) > 0 else 0
                total += cnt
                c = ws.cell(r, col + j, cnt if cnt > 0 else 0)
                style_cell(c, NORMAL_FONT, alignment=CENTER, border=THIN_BORDER)
            c = ws.cell(r, col + 3, total)
            style_cell(c, BOLD_FONT, alignment=CENTER, border=THIN_BORDER)
            col += 4

    # TOTAL row
    total_row = 3 + len(employees)
    ws.cell(total_row, 1, "TOTAL")
    style_cell(ws.cell(total_row, 1), BOLD_FONT, alignment=LEFT, border=THIN_BORDER)
    col = 2
    for m in months:
        for j, cat in enumerate(CAT_ORDER):
            col_sum = sum(
                int(pivot[(pivot["emp_name"] == emp) & (pivot["month_dt"] == m) & (pivot["category"] == cat)]["count"].values[0])
                if len(pivot[(pivot["emp_name"] == emp) & (pivot["month_dt"] == m) & (pivot["category"] == cat)]) > 0 else 0
                for emp in employees
            )
            c = ws.cell(total_row, col + j, col_sum)
            style_cell(c, BOLD_FONT, alignment=CENTER, border=THIN_BORDER)
        month_total = sum(
            int(pivot[(pivot["emp_name"] == emp) & (pivot["month_dt"] == m)]["count"].sum())
            if len(pivot[(pivot["emp_name"] == emp) & (pivot["month_dt"] == m)]) > 0 else 0
            for emp in employees
        )
        c = ws.cell(total_row, col + 3, month_total)
        style_cell(c, BOLD_FONT, alignment=CENTER, border=THIN_BORDER)
        col += 4


def build_daily_records(wb, df):
    ws = wb.create_sheet("Daily Late Records")

    headers = ["Employee Code", "Employee Name", "Date", "Day", "Clock In", "Late (mins)", "Category"]
    widths = [14, 28, 14, 8, 10, 12, 16]
    for j, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, j, h)
        style_cell(c, WHITE_FONT, DARK_BLUE_FILL, CENTER, THIN_BORDER)
        ws.column_dimensions[get_column_letter(j)].width = w

    for i, (_, row) in enumerate(df.iterrows(), 2):
        ws.cell(i, 1, row["emp_code"]).alignment = CENTER
        ws.cell(i, 2, row["emp_name"]).alignment = LEFT
        ws.cell(i, 3, row["date"].strftime("%d-%b-%Y")).alignment = CENTER
        ws.cell(i, 4, row["day_name"]).alignment = CENTER
        ws.cell(i, 5, row["clock_in"]).alignment = CENTER

        mins_cell = ws.cell(i, 6, row["late_mins"])
        cat_cell = ws.cell(i, 7, row["category"])
        fill = cat_fill(row["category"])
        font = cat_font(row["category"])
        style_cell(mins_cell, font, fill, CENTER, THIN_BORDER)
        style_cell(cat_cell, font, fill, CENTER, THIN_BORDER)

        for j in range(1, 6):
            ws.cell(i, j).border = THIN_BORDER
            ws.cell(i, j).font = NORMAL_FONT


def _write_employee_block(ws, row_start, emp_code, emp_name, emp_df, months, month_labels, max_col=7):
    """Write a single employee block (used by Employee Detail and Top N sheets).
    Returns the next available row."""
    r = row_start

    # Title row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    c = ws.cell(r, 1, f"{emp_code} - {emp_name}")
    style_cell(c, WHITE_FONT_12, DARK_BLUE_FILL, LEFT)
    r += 1

    # Monthly summary sub-header
    sum_headers = ["Month", "Within 15 mins", "16-30 min", "Over 30 min", "Total Late Days"]
    sum_fills = [DARK_BLUE_FILL, GREEN_FILL, YELLOW_FILL, RED_FILL, GRAY_FILL]
    sum_fonts = [WHITE_FONT, BOLD_FONT, BOLD_FONT, BOLD_FONT, BOLD_FONT]
    for j, (h, f, fn) in enumerate(zip(sum_headers, sum_fills, sum_fonts), 1):
        c = ws.cell(r, j, h)
        style_cell(c, fn, f, CENTER, THIN_BORDER)
    r += 1

    # Monthly rows
    cat_pivot = emp_df.groupby(["month_dt", "category"]).size().reset_index(name="count")
    total_all_cats = [0, 0, 0, 0]  # within15, 16-30, over30, total
    for m, label in zip(months, month_labels):
        ws.cell(r, 1, label)
        style_cell(ws.cell(r, 1), NORMAL_FONT, alignment=CENTER, border=THIN_BORDER)
        row_total = 0
        for j, cat in enumerate(CAT_ORDER, 2):
            val = cat_pivot[(cat_pivot["month_dt"] == m) & (cat_pivot["category"] == cat)]
            cnt = int(val["count"].values[0]) if len(val) > 0 else 0
            row_total += cnt
            total_all_cats[j - 2] += cnt
            c = ws.cell(r, j, cnt)
            style_cell(c, NORMAL_FONT, alignment=CENTER, border=THIN_BORDER)
        total_all_cats[3] += row_total
        c = ws.cell(r, 5, row_total)
        style_cell(c, NORMAL_FONT, alignment=CENTER, border=THIN_BORDER)
        r += 1

    # Totals row for monthly summary
    ws.cell(r, 1, "Total")
    style_cell(ws.cell(r, 1), BOLD_FONT, alignment=CENTER, border=THIN_BORDER)
    for j in range(4):
        c = ws.cell(r, j + 2, total_all_cats[j])
        style_cell(c, BOLD_FONT, alignment=CENTER, border=THIN_BORDER)

    # Total late mins
    total_mins = int(emp_df["late_mins"].sum())
    if max_col >= 10:
        ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=10)
        c = ws.cell(r, 8, f"Total Late {total_mins} mins")
        style_cell(c, BOLD_FONT, alignment=CENTER)
    r += 2

    return r, total_all_cats[3], total_mins


def build_employee_detail(wb, df, months, month_labels):
    ws = wb.create_sheet("Employee Detail")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 16

    employees = sorted(df[["emp_code", "emp_name"]].drop_duplicates().values.tolist())
    r = 1
    for emp_code, emp_name in employees:
        emp_df = df[df["emp_code"] == emp_code]
        r, _, _ = _write_employee_block(ws, r, emp_code, emp_name, emp_df, months, month_labels)

        # Daily detail sub-header
        detail_headers = ["Date", "Day", "Clock In", "Late (mins)", "Category"]
        for j, h in enumerate(zip(detail_headers, [DARK_BLUE_FILL]*5), 1):
            c = ws.cell(r, j, h[0])
            style_cell(c, WHITE_FONT, DARK_BLUE_FILL, CENTER, THIN_BORDER)
        r += 1

        for _, row in emp_df.iterrows():
            ws.cell(r, 1, row["date"].strftime("%d-%b-%Y")).alignment = CENTER
            ws.cell(r, 2, row["day_name"]).alignment = CENTER
            ws.cell(r, 3, row["clock_in"]).alignment = CENTER
            mins_c = ws.cell(r, 4, row["late_mins"])
            cat_c = ws.cell(r, 5, row["category"])
            fill = cat_fill(row["category"])
            font = cat_font(row["category"])
            style_cell(mins_c, font, fill, CENTER, THIN_BORDER)
            style_cell(cat_c, font, fill, CENTER, THIN_BORDER)
            for j in range(1, 4):
                ws.cell(r, j).border = THIN_BORDER
                ws.cell(r, j).font = NORMAL_FONT
            r += 1

        r += 2  # blank rows between employees


def build_top_n_sheet(wb, df, months, month_labels, top_n):
    # Determine date range for sheet name
    first_month = month_labels[0].replace(" ", "")
    last_month = month_labels[-1].replace(" ", "")
    sheet_name = f"{first_month}-{last_month} Top {top_n}"
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    ws = wb.create_sheet(sheet_name)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 18

    # Find top N by total late count
    emp_counts = df.groupby(["emp_code", "emp_name"]).size().reset_index(name="total_count")
    emp_counts = emp_counts.sort_values("total_count", ascending=False).head(top_n)

    r = 1
    for _, emp_row in emp_counts.iterrows():
        emp_code = emp_row["emp_code"]
        emp_name = emp_row["emp_name"]
        emp_df = df[df["emp_code"] == emp_code]
        r, _, _ = _write_employee_block(ws, r, emp_code, emp_name, emp_df, months, month_labels, max_col=10)
        r += 1  # extra blank row between employees


# ── Main entry point ──
def generate_lateness_report(input_path, output_path, top_n=10, status_cb=None):
    def status(msg):
        if status_cb:
            status_cb(msg)

    status("Loading attendance data...")
    df = load_attendance(input_path)

    if df.empty:
        raise ValueError("No late records found in the attendance file.")

    months, month_labels = get_sorted_months(df)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    status("Building Summary of Late in Mins...")
    build_summary_mins(wb, df, months, month_labels, top_n)

    status("Building Summary of Late in Freq...")
    build_summary_freq(wb, df, months, month_labels, top_n)

    status("Building Monthly Summary...")
    build_monthly_summary(wb, df, months, month_labels)

    status("Building Daily Late Records...")
    build_daily_records(wb, df)

    status("Building Employee Detail...")
    build_employee_detail(wb, df, months, month_labels)

    status(f"Building Top {top_n} sheet...")
    build_top_n_sheet(wb, df, months, month_labels, top_n)

    status("Saving report...")
    wb.save(output_path)
    status("Done!")
